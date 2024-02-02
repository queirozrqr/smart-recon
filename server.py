import os
import http.server
import socketserver

from http import HTTPStatus
from operator import truediv
from ast import Try
import json, requests,  time,  sys, os
from datetime import datetime
from urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        self.send_response(HTTPStatus.OK)
        self.end_headers()

        global sheet_programa

        use_bbrf = 0
        use_dirs = 1
        debug = 0
        BASE_DIR="/bugbounty/programs/intigriti"
        RECON_DIR="/bugbounty/recon/intigriti"
        API_URL="https://api.intigriti.com/external/researcher/v1"

        API_PROD="https://login.intigriti.com/connect/authorize"
        API_UAT="https://login-uat.intigriti.com/connect/authorize"

        INT_PROGS ="programs"

        TOKEN="13AC2648DCEF4536D0714619D5B6F5FE313ACEF7D75C0C2CCF35E4982E0CDBF6-1"
        if(use_bbrf):
            from bbrf import BBRFClient as bbrf
            conf = {
            "username": "admin",
            "password": "admin",
            "couchdb": "https://bbrf.r1ck.xyz/bbrf",
            "ignore_ssl_errors": True,
            "program": "datadome-bot-bounty"
            }
            bbrf_programs = bbrf('programs --show-disabled --show-empty-scope', conf).run()

        if(use_dirs):
            import openpyxl
            BASE_DIR="/bugbounty/programs/intigriti/"

            wb = openpyxl.Workbook()
            sheet_programa = wb.active
            sheet_programa.title = "Programas"
            sheet_programa.cell(row=1, column=1).value = 'Plataforma'
            sheet_programa.cell(row=1, column=2).value = 'Nome Empresa'
            sheet_programa.cell(row=1, column=3).value = 'ID Empresa'
            sheet_programa.cell(row=1, column=4).value = 'ID Programa'
            sheet_programa.cell(row=1, column=5).value = 'Bounty'
            sheet_programa.cell(row=1, column=6).value = 'Privado'
            sheet_programa.cell(row=1, column=7).value = 'Estado'
            sheet_programa.cell(row=1, column=8).value = 'VDP'
            sheet_programa.cell(row=1, column=9).value = 'Habilitado'
            sheet_programa.cell(row=1, column=10).value = 'Intigriti triaged'
            sheet_programa.cell(row=1, column=11).value = 'reports_count'
            sheet_programa.cell(row=1, column=12).value = 'bounty_reward_max'
            sheet_programa.cell(row=1, column=13).value = 'bounty_reward_min'
            sheet_programa.cell(row=1, column=14).value = 'currency'
            sheet_programa.cell(row=1, column=15).value = 'Credentials'
            sheet_programa.cell(row=1, column=16).value = 'UpdatedAt'

            sheet_inscope = wb.create_sheet("Scope_IN")
            sheet_inscope.cell(row=1, column=1).value = 'Plataforma'
            sheet_inscope.cell(row=1, column=2).value = 'Nome Empresa'
            sheet_inscope.cell(row=1, column=3).value = 'ID Empresa'
            sheet_inscope.cell(row=1, column=4).value = 'Tipo'
            sheet_inscope.cell(row=1, column=5).value = 'Endpoint'
            sheet_inscope.cell(row=1, column=6).value = 'BusinessImpact'
            sheet_inscope.cell(row=1, column=7).value = 'Description'

            sheet_outscope = wb.create_sheet("Scope_OUT")
            sheet_outscope.cell(row=1, column=1).value = 'Plataforma'
            sheet_outscope.cell(row=1, column=2).value = 'Nome Empresa'
            sheet_outscope.cell(row=1, column=3).value = 'ID Empresa'
            sheet_outscope.cell(row=1, column=4).value = 'Content'
            sheet_outscope.cell(row=1, column=5).value = 'Endpoint'
            sheet_outscope.cell(row=1, column=6).value = 'BusinessImpact'
            sheet_outscope.cell(row=1, column=7).value = 'Description'
            
            #wb.save(BASE_DIR+'intigriti.xlsx')
        limit=2
        pag=0
        cont=1

        MyRequestHeaders = {'Authorization': 'Bearer 13AC2648DCEF4536D0714619D5B6F5FE313ACEF7D75C0C2CCF35E4982E0CDBF6-1', 'Content-Type': 'application/json'}
        ProgramsResponse = requests.get(API_URL+'/programs?statusId=3&typeId=1&limit=500&offset=1', headers=MyRequestHeaders, verify=False)
        ProgramsJson = json.loads(ProgramsResponse.content)
        items = ProgramsJson['maxCount']
        pages = items / 500
        if(pages<1): pages=1
        sheet_programa_linha = 1
        sheet_inscope_linha = 1
        sheet_outscope_linha = 1

        for page in range(0, int(pages)):
            MyRequestHeaders = {'Authorization': 'Bearer 13AC2648DCEF4536D0714619D5B6F5FE313ACEF7D75C0C2CCF35E4982E0CDBF6-1', 'Content-Type': 'application/json'}
            ProgramsResponse = requests.get(API_URL+'/programs?statusId=3&typeId=1&limit=500&offset='+str(cont), headers=MyRequestHeaders, verify=False)
            ProgramsJson = json.loads(ProgramsResponse.content)
            limit = ProgramsJson['maxCount']    
            types=[]
            for i in ProgramsJson['records']:
                cont+=1
                if(use_dirs):
                    sheet_programa_linha=sheet_programa_linha+1
                private=i['confidentialityLevel']['value']=="InviteOnly"
                status=i['status']['value']
                bounty=i['type']['value']
                handle=i['handle']
                programId=i['id']
                name=i['name']

                try:
                    currency=i['minBounty']['currency']
                    bounty_reward_min=i['minBounty']['value']
                    bounty_reward_max=i['maxBounty']['value']

                except:
                    currency=''
                    bounty_reward_min=''
                    bounty_reward_max=''
                        
                if(use_bbrf):
                    try:
                        if(handle in bbrf_programs):
                            bbrf_cmd='program update ' + handle + ' -t plataform:intigriti -t Handle:'+ handle + ' -t companyHandle:'+ companyHandle + ' -t private:'+ str(private)
                            #print(bbrf_cmd)
                            bbrf(bbrf_cmd, conf).run()
                        else:
                            bbrf_cmd='new ' + handle + ' -t plataform:intigriti -t Handle:'+ handle + ' -t companyHandle:'+ companyHandle + ' -t private:'+ str(private)
                            #print(bbrf_cmd)
                            bbrf(bbrf_cmd, conf).run()
                    except:
                        continue
                # if(i['tacRequired']):
                #     try:
                #         ProgramTesteResponse = requests.get(API_URL+'/programs/'+companyHandle+'/'+handle,  headers=MyRequestHeaders, verify=False)
                #         ProgramTesteJson = json.loads(ProgramTesteResponse.content)
                #         if(ProgramTesteResponse.status_code!=200):
                #             ProgramAccept = requests.put(API_URL+'/programs/tac/' + programId + '/accept',  headers=MyRequestHeaders, verify=False)
                #             time.sleep(5)

                #     except:
                #         print('line 155', i)
                #         print(ProgramResponse.status_code)
                #         time.sleep(5)
                
                ProgramResponse = requests.get(API_URL+'/programs/'+programId,  headers=MyRequestHeaders, verify=False)
                if(ProgramResponse.status_code == 200):
                    ProgramJson = json.loads(ProgramResponse.content)
                    lastUpdatedAt=ProgramJson['domains']['createdAt']
                    createdAt = datetime.fromtimestamp(lastUpdatedAt, tz=None)

                    if(use_dirs):
                        sheet_programa.cell(row=sheet_programa_linha, column=1).value = "INTIGRITI"
                        sheet_programa.cell(row=sheet_programa_linha, column=2).value = name
                        sheet_programa.cell(row=sheet_programa_linha, column=4).value = handle
                        sheet_programa.cell(row=sheet_programa_linha, column=5).value = bounty
                        sheet_programa.cell(row=sheet_programa_linha, column=6).value = str(private)
                        sheet_programa.cell(row=sheet_programa_linha, column=7).value = status
                        sheet_programa.cell(row=sheet_programa_linha, column=8).value = bounty
                        sheet_programa.cell(row=sheet_programa_linha, column=12).value = bounty_reward_max
                        sheet_programa.cell(row=sheet_programa_linha, column=13).value = bounty_reward_min
                        sheet_programa.cell(row=sheet_programa_linha, column=14).value = currency
                        if(lastUpdatedAt!="ERRO"):
                            sheet_programa.cell(row=sheet_programa_linha, column=16).value = datetime.fromtimestamp(lastUpdatedAt, tz=None)
                        
                    URL=''
                    APPLE_STORE_APP_ID=''
                    GOOGLE_PLAY_APP_ID=''
                    CIDR=''
                    TYPE5=''
                    API=''

                    for d in ProgramJson['domains']['content']:
                        description=d['description']
                        tier=d['tier']['value']
                        endpoint=d['endpoint']
                        tipo=d['type']['id']

                        try:
                            businessImpact=d['businessImpact']
                        except:
                            businessImpact=''
                                        
                        if(tipo==1):
                            URL += endpoint + ' '
                            strtipo='URL'

                        elif(tipo==2):
                            GOOGLE_PLAY_APP_ID += endpoint + ' '
                            strtipo="GOOGLE_PLAY_APP"
                            if(use_bbrf):
                                bbrf_cmd='program update '+ handle + ' -t mobile_android:'+ endpoint + ' --append-tags'
                                #print(bbrf_cmd)
                                bbrf(bbrf_cmd, conf).run()

                        elif(tipo==3):
                            APPLE_STORE_APP_ID += endpoint + ' '
                            strtipo="APPLE_STORE_APP"
                            if(use_bbrf):
                                bbrf_cmd='program update '+ handle + ' -t mobile_ios:'+ endpoint + ' --append-tags'
                                #print(bbrf_cmd)
                                bbrf(bbrf_cmd, conf).run()

                        elif(tipo==4):
                            CIDR += endpoint + ' '
                            strtipo="CIDR"

                        elif(tipo==5):
                            TYPE5 += endpoint + ' '
                            strtipo="TIPO 5"

                        elif(tipo==6):
                            API += endpoint + ' '
                            strtipo="API"

                        if(use_dirs):
                            if(tier!='Out Of Scope'):
                                sheet_inscope_linha=sheet_inscope_linha+1
                                sheet_inscope.cell(row=sheet_inscope_linha, column=1).value = 'INTIGRITI'
                                sheet_inscope.cell(row=sheet_inscope_linha, column=2).value = name
                                sheet_inscope.cell(row=sheet_inscope_linha, column=3).value = handle
                                sheet_inscope.cell(row=sheet_inscope_linha, column=4).value = strtipo
                                sheet_inscope.cell(row=sheet_inscope_linha, column=5).value = endpoint
                                sheet_inscope.cell(row=sheet_inscope_linha, column=6).value = businessImpact
                                sheet_inscope.cell(row=sheet_inscope_linha, column=7).value = description
                                #sheet_inscope.cell(row=sheet_inscope_linha, column=8).value = str(d)
                            else:
                                sheet_outscope_linha=sheet_outscope_linha+1
                                sheet_outscope.cell(row=sheet_outscope_linha, column=1).value = "INTIGRITI"
                                sheet_outscope.cell(row=sheet_outscope_linha, column=2).value = name
                                sheet_outscope.cell(row=sheet_outscope_linha, column=3).value = handle
                                sheet_outscope.cell(row=sheet_outscope_linha, column=4).value = strtipo
                                sheet_outscope.cell(row=sheet_outscope_linha, column=5).value = endpoint
                                sheet_outscope.cell(row=sheet_outscope_linha, column=6).value = businessImpact
                                sheet_outscope.cell(row=sheet_outscope_linha, column=7).value = description
                                #sheet_outscope.cell(row=sheet_outscope_linha, column=8).value = str(d)


                    if(use_bbrf):
                        if(len(URL)>0):
                            #print("URLs In Scope:", URL)
                            try:
                                bbrf('inscope add ' + URL + ' -p ' + handle , conf).run()
                                bbrf('domain add ' + URL + ' -p ' + handle , conf).run()
                            except:
                                print("Erro on domain/scope add: ", handle, URL) 
                                continue
                    if(use_dirs):
                        wb.save(BASE_DIR+'intigriti.xlsx')
                else:
                    if(use_dirs):
                        sheet_programa.cell(row=sheet_programa_linha, column=1).value = "INTIGRITI"
                        sheet_programa.cell(row=sheet_programa_linha, column=2).value = name
                        sheet_programa.cell(row=sheet_programa_linha, column=4).value = handle
                        sheet_programa.cell(row=sheet_programa_linha, column=5).value = 'SEM AUTORIZAÇÃO'

        if(use_dirs):
            wb.save(BASE_DIR+'intigriti.xlsx')

        
        msg = 'Hello! you requested %s' % (self.path)
        self.wfile.write(msg.encode())


port = int(os.getenv('PORT', 80))
print('Listening on port %s' % (port))
httpd = socketserver.TCPServer(('', port), Handler)
httpd.serve_forever()
